"""Deterministic Excel export of a valuation run. Pure read of the run record:
regroups the persisted per-well cashflow schedule into static driver columns,
aggregates across statuses, then writes a live two-sheet Excel model — a Summary
and a single editable Cashflow statement (econ as formulas, volumes frozen,
prices + assumptions exposed as editable inputs). No discounting: PV is left to
the user. No economics recompute.
See docs/superpowers/plans/2026-06-25-valuation-excel-export.md.
"""
import io
import re
from datetime import date

from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

from server.valuation import config

STATUS_ORDER = ["PDP", "DUC", "PUD"]


class ExportError(Exception):
    """Run record is missing data the export needs (e.g. per-well schedule)."""


def _online_offset(oil: list, gas: list) -> int:
    """First month index with nonzero oil or gas. len(oil) if never online."""
    for i, (o, g) in enumerate(zip(oil, gas)):
        if o > 0 or g > 0:
            return i
    return len(oil)


def build_status_drivers(economics: dict, wells: dict) -> dict:
    """Per-status static driver columns the workbook formulas multiply against.

    Returns {status_code: {gross_oil, gross_gas, active_count, new_wells}} for
    every status code in STATUS_ORDER that has >=1 well. Lists are length
    horizon_months. Raises ExportError when the per-well schedule was omitted
    (deals above the 200-well audit cap)."""
    horizon = int(economics["horizon_months"])
    sched = economics["schedule"]
    by_well = sched.get("by_well")
    if not by_well:
        raise ExportError(
            "economics.schedule.by_well is absent (deal exceeds the per-well "
            "audit cap); Excel export needs per-well rows")
    statuses = wells.get("statuses", {})

    out: dict[str, dict] = {}
    for api, cols in by_well.items():
        code = config.status_code(statuses.get(api))
        d = out.setdefault(code, {
            "gross_oil": [0.0] * horizon, "gross_gas": [0.0] * horizon,
            "active_count": [0] * horizon, "new_wells": [0] * horizon,
        })
        oil, gas, capex = cols["oil_bbl"], cols["gas_mcf"], cols["capex"]
        offset = _online_offset(oil, gas)
        for m in range(horizon):
            d["gross_oil"][m] += float(oil[m])
            d["gross_gas"][m] += float(gas[m])
            if m >= offset:
                d["active_count"][m] += 1
            if capex[m] > 0:
                d["new_wells"][m] += 1
    # Stable, deal-sheet order; drop empty buckets.
    return {code: out[code] for code in STATUS_ORDER if code in out}


def read_assumptions(economics: dict) -> dict:
    """Extract every economic assumption needed by the oracle from the persisted
    economics stage. ``oil_price``/``gas_price`` are month-by-month vectors read
    from ``price_path``; legacy runs without it fall back to the flat scalar
    repeated ``horizon`` times."""
    horizon = int(economics["horizon_months"])
    inputs = economics.get("inputs", {})
    interest = economics.get("interest", {})
    costs = economics.get("cost_inputs", {})
    pp = economics.get("price_path")
    if pp and pp.get("oil") and pp.get("gas"):
        oil_price = [float(v) for v in pp["oil"]]
        gas_price = [float(v) for v in pp["gas"]]
    else:                                            # legacy run: flat fallback
        oil_price = [float(inputs.get("oil_price", 0.0))] * horizon
        gas_price = [float(inputs.get("gas_price", 0.0))] * horizon
    return {
        "interest_type": interest.get("interest_type"),
        "wi": interest.get("wi_pct"), "nri": interest.get("nri_pct"),
        "decimal": interest.get("decimal"),
        "oil_diff": float(inputs.get("oil_diff", 0.0)),
        "gas_diff": float(inputs.get("gas_diff", 0.0)),
        "tax_pct": float(inputs.get("tax_pct", 0.0)),
        "gpt_pct": float(inputs.get("gpt_pct", 0.0)),
        "opex_well_month": float(costs.get("opex_per_well_month", 0.0)),
        "opex_bbl": float(costs.get("opex_per_bbl", 0.0)),
        "capex_per_well": float(costs.get("capex_per_well", 0.0)),
        "oil_price": oil_price, "gas_price": gas_price,
        "rate_centers": economics.get("rate_centers", {}),
    }


def status_net_cashflow(drivers_one: dict, a: dict) -> list[float]:
    """Monthly net cashflow for one status bucket. Mirrors econ.cashflow_components."""
    horizon = len(drivers_one["gross_oil"])
    out = [0.0] * horizon
    wi, nri, dec = a["wi"], a["nri"], a["decimal"]
    for m in range(horizon):
        oil, gas = drivers_one["gross_oil"][m], drivers_one["gross_gas"][m]
        gross_rev = oil * (a["oil_price"][m] - a["oil_diff"]) + \
            gas * (a["gas_price"][m] - a["gas_diff"])
        if a["interest_type"] == "wi":
            net_rev = nri * gross_rev
            sev_tax = a["tax_pct"] * wi * gross_rev
            gpt = a["gpt_pct"] * wi * gross_rev
            opex = wi * (a["opex_well_month"] * drivers_one["active_count"][m]
                         + a["opex_bbl"] * oil)
            capex = wi * a["capex_per_well"] * drivers_one["new_wells"][m]
        else:                                        # minerals
            net_rev = dec * gross_rev
            sev_tax = a["tax_pct"] * dec * gross_rev
            gpt = opex = capex = 0.0
        out[m] = net_rev - sev_tax - gpt - opex - capex
    return out


def npv_monthly(cashflow: list[float], annual_rate: float) -> float:
    """Ordinary-annuity NPV, t=1..n. Equals econ.npv and Excel =NPV(rate/12, range)."""
    if annual_rate == 0.0:
        return float(sum(cashflow))
    monthly = annual_rate / 12.0
    return float(sum(cf / (1.0 + monthly) ** (i + 1) for i, cf in enumerate(cashflow)))


def reconcile_total_pv(economics: dict, wells: dict) -> float:
    """Sum NPV over all status buckets, reproducing ``economics["npv_at_centers"]["total"]``
    within $1 using the mirrored oracle algebra (independent of econ.py)."""
    drivers = build_status_drivers(economics, wells)
    a = read_assumptions(economics)
    total = 0.0
    for code, d in drivers.items():
        total += npv_monthly(status_net_cashflow(d, a), float(a["rate_centers"][code]))
    return total


# ── Workbook assembly ─────────────────────────────────────────────────────────

_TITLE_ROW = 1               # Cashflow sheet title
_INPUT_ROW0 = 4              # first scalar-input row (band sits rows 4-8)
_TABLE_HDR_ROW = 10          # Cashflow monthly-table header row
_FIRST_DATA_ROW = _TABLE_HDR_ROW + 1

# Styling: amber = "edit me", graphite = header band (matches the app chrome).
_INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
_HDR_FILL = PatternFill("solid", fgColor="15181B")
_THIN = Side(style="thin", color="D9D9D9")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def export_filename(facts: dict, run_id: str) -> str:
    """Filename for the exported workbook, e.g. ``CrudeCode_Valuation_Reeves-County_abcdef12.xlsx``."""
    slug = re.sub(r"[^A-Za-z0-9]+", "-", (facts.get("area") or "deal")).strip("-")
    return f"CrudeCode_Valuation_{slug}_{run_id[:8]}.xlsx"


def _months(origin: str, horizon: int) -> list:
    """List of ``date`` objects, one per forecast month starting at ``origin``."""
    cur = date.fromisoformat(origin)
    out = []
    for _ in range(horizon):
        out.append(cur)
        cur = cur + relativedelta(months=1)
    return out


def _name(wb, name: str, sheet: str, cell: str) -> None:
    """Add a workbook-scoped named range pointing at a single cell."""
    ref = f"{quote_sheetname(sheet)}!${cell[0]}${cell[1:]}"
    wb.defined_names.add(DefinedName(name, attr_text=ref))


def _aggregate_drivers(drivers: dict, horizon: int) -> dict:
    """Sum the per-status driver columns into one aggregate set. Valid because
    every downstream cashflow term (rev, tax, opex, capex) is linear in these
    columns, so summing the drivers then applying the algebra once equals summing
    the per-status cashflows."""
    agg = {"gross_oil": [0.0] * horizon, "gross_gas": [0.0] * horizon,
           "active_count": [0] * horizon, "new_wells": [0] * horizon}
    for d in drivers.values():
        for m in range(horizon):
            agg["gross_oil"][m] += d["gross_oil"][m]
            agg["gross_gas"][m] += d["gross_gas"][m]
            agg["active_count"][m] += d["active_count"][m]
            agg["new_wells"][m] += d["new_wells"][m]
    return agg


def _style_input(cell, fmt: str) -> None:
    """Mark a cell as an editable input: amber fill, border, number format."""
    cell.fill = _INPUT_FILL
    cell.border = _BORDER
    cell.number_format = fmt


def _build_inputs_band(wb, ws, a: dict, is_wi: bool) -> None:
    """Title + scalar-assumption inputs along the top of the Cashflow sheet.
    Each value cell is amber and exposed as a workbook-scoped named range so the
    monthly formulas can reference it by name regardless of its address."""
    title = ws.cell(row=_TITLE_ROW, column=1,
                    value="CRUDE CODE — EDITABLE INPUTS   ▸ change the amber cells")
    title.font = Font(bold=True, size=13, color="15181B")
    note = ws.cell(row=_TITLE_ROW + 1, column=1,
                   value="Amber cells are yours to edit — the model recomputes automatically. "
                         "Prices are in the table below, one per month.")
    note.font = Font(italic=True, size=9, color="7A7870")

    left = [
        ("Oil differential ($/bbl)",   a["oil_diff"],        "oil_diff",   "#,##0.00"),
        ("Gas differential ($/mmbtu)", a["gas_diff"],        "gas_diff",   "#,##0.00"),
        ("Opex $/well/month",          a["opex_well_month"], "opex_well",  "$#,##0"),
        ("Opex $/bbl",                 a["opex_bbl"],        "opex_bbl",   "$#,##0.00"),
        ("Capex $/well",               a["capex_per_well"],  "capex_well", "$#,##0"),
    ]
    right = [
        ("Severance tax (frac)", a["tax_pct"], "tax_pct", "0.0%"),
        ("GPT (frac)",           a["gpt_pct"], "gpt_pct", "0.0%"),
    ]
    if is_wi:
        right += [("Working interest (frac)",     a["wi"],  "wi_pct",  "0.0%"),
                  ("Net revenue interest (frac)", a["nri"], "nri_pct", "0.0%")]
    else:
        right += [("Mineral decimal interest", a["decimal"], "dec_int", "0.00000000")]

    for col_label, col_val, items in (("A", "B", left), ("D", "E", right)):
        for i, (label, val, nm, fmt) in enumerate(items):
            r = _INPUT_ROW0 + i
            lbl = ws[f"{col_label}{r}"]
            lbl.value = label
            lbl.font = Font(bold=True)
            cell = ws[f"{col_val}{r}"]
            cell.value = val
            _style_input(cell, fmt)
            _name(wb, nm, ws.title, f"{col_val}{r}")


def _build_cashflow_table(ws, a: dict, agg: dict, months: list, horizon: int,
                          is_wi: bool) -> tuple[int, int]:
    """The aggregate monthly cashflow table. Prices (cols F/G) are editable inputs;
    everything right of them is formulas. Returns (first_row, last_row)."""
    hdr = ["Month", "Gross Oil (bbl)", "Gross Gas (mcf)", "Active wells",
           "New wells", "Oil $/bbl", "Gas $/mmbtu", "Gross Rev", "Net Rev",
           "Sev Tax", "GPT", "Opex", "Capex", "Net Cashflow", "Cumulative NCF"]
    for c, label in enumerate(hdr, start=1):
        cell = ws.cell(row=_TABLE_HDR_ROW, column=c, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    first = _FIRST_DATA_ROW
    for m in range(horizon):
        r = first + m
        ws[f"A{r}"] = months[m]
        ws[f"A{r}"].number_format = "yyyy-mm"
        ws[f"B{r}"] = round(agg["gross_oil"][m], 2)
        ws[f"C{r}"] = round(agg["gross_gas"][m], 2)
        ws[f"D{r}"] = agg["active_count"][m]
        ws[f"E{r}"] = agg["new_wells"][m]
        _style_input(ws[f"F{r}"], "#,##0.00")
        ws[f"F{r}"].value = round(a["oil_price"][m], 2)
        _style_input(ws[f"G{r}"], "#,##0.00")
        ws[f"G{r}"].value = round(a["gas_price"][m], 2)
        ws[f"H{r}"] = f"=B{r}*(F{r}-oil_diff)+C{r}*(G{r}-gas_diff)"
        if is_wi:
            ws[f"I{r}"] = f"=nri_pct*H{r}"
            ws[f"J{r}"] = f"=tax_pct*wi_pct*H{r}"
            ws[f"K{r}"] = f"=gpt_pct*wi_pct*H{r}"
            ws[f"L{r}"] = f"=wi_pct*(opex_well*D{r}+opex_bbl*B{r})"
            ws[f"M{r}"] = f"=wi_pct*capex_well*E{r}"
        else:
            ws[f"I{r}"] = f"=dec_int*H{r}"
            ws[f"J{r}"] = f"=tax_pct*dec_int*H{r}"
            ws[f"K{r}"] = 0
            ws[f"L{r}"] = 0
            ws[f"M{r}"] = 0
        ws[f"N{r}"] = f"=I{r}-J{r}-K{r}-L{r}-M{r}"
        ws[f"O{r}"] = f"=N{r}" if m == 0 else f"=O{r - 1}+N{r}"
        ws[f"B{r}"].number_format = "#,##0"
        ws[f"C{r}"].number_format = "#,##0"
        for col in "HIJKLMNO":
            ws[f"{col}{r}"].number_format = "$#,##0"

    ws.column_dimensions["A"].width = 10
    for col in "BCHIJKLMNO":
        ws.column_dimensions[col].width = 13
    return first, first + horizon - 1


def _build_summary(wb, facts: dict, run_id: str, well_count: int,
                   first: int, last: int) -> None:
    """Deal facts + an undiscounted totals roll-up referencing the Cashflow sheet.
    No PV line — discounting is the user's job."""
    summ = wb.create_sheet("Summary", index=0)
    summ["A1"] = "Crude Code — Valuation"
    summ["A1"].font = Font(bold=True, size=14)
    facts_rows = [
        ("Deal type", facts.get("deal_type")),
        ("Interest",  facts.get("interest")),
        ("Operator",  facts.get("operator")),
        ("Area",      facts.get("area")),
        ("Wells",     well_count),
        ("Run ID",    run_id),
    ]
    for i, (label, val) in enumerate(facts_rows, start=2):
        summ[f"A{i}"] = label
        summ[f"A{i}"].font = Font(bold=True)
        summ[f"B{i}"] = val

    base = 2 + len(facts_rows) + 1
    summ[f"A{base}"] = "Undiscounted totals"
    summ[f"A{base}"].font = Font(bold=True, size=12)
    totals = [
        ("Total Gross Oil (bbl)", f"=SUM(Cashflow!B{first}:B{last})", "#,##0"),
        ("Total Gross Gas (mcf)", f"=SUM(Cashflow!C{first}:C{last})", "#,##0"),
        ("Total Net Cashflow",    f"=SUM(Cashflow!N{first}:N{last})", "$#,##0"),
    ]
    for i, (label, formula, fmt) in enumerate(totals, start=base + 1):
        summ[f"A{i}"] = label
        summ[f"A{i}"].font = Font(bold=True)
        summ[f"B{i}"] = formula
        summ[f"B{i}"].number_format = fmt
    summ[f"A{base + len(totals) + 2}"] = (
        "Discounting is intentionally left to you — apply your own rate to the "
        "monthly Net Cashflow on the Cashflow sheet.")
    summ.column_dimensions["A"].width = 22
    summ.column_dimensions["B"].width = 30


def build_workbook_bytes(run_id: str, economics: dict, wells: dict, facts: dict) -> bytes:
    """Assemble and return the live Excel model as raw bytes: a Summary sheet and
    a single editable Cashflow statement aggregated across statuses.

    Volumes are frozen static; price/cost/interest inputs are editable workbook-scoped
    named ranges; revenue → net-cashflow are Excel formula strings that encode the
    same algebra as ``status_net_cashflow``. No discounting — PV is left to the user.

    Raises ``ExportError`` when ``economics.schedule.by_well`` is absent."""
    drivers = build_status_drivers(economics, wells)        # raises ExportError if no by_well
    a = read_assumptions(economics)
    horizon = int(economics["horizon_months"])
    months = _months(economics["schedule"]["origin"], horizon)
    agg = _aggregate_drivers(drivers, horizon)
    is_wi = a["interest_type"] == "wi"
    well_count = len(economics["schedule"]["by_well"])

    wb = Workbook()
    cf = wb.active
    cf.title = "Cashflow"
    _build_inputs_band(wb, cf, a, is_wi)
    first, last = _build_cashflow_table(cf, a, agg, months, horizon, is_wi)
    _build_summary(wb, facts, run_id, well_count, first, last)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
