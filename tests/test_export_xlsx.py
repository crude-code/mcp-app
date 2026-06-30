import io

import pytest
from openpyxl import load_workbook

from server.valuation import export_xlsx as ex
from server.valuation import orchestrator as orch
from server.valuation.types import DeclineCurve, ForecastProvenance


def _mk_curve(stream):
    c = DeclineCurve(qi_peak=600.0, di=0.13, b=0.8, terminal_di_monthly=0.005,
                     switch_month_from_peak=float("inf"), stream=stream,
                     provenance=ForecastProvenance(source="self", strategy="pdp"))
    return orch._serialize_curve(c)


def _engine_minerals_run():
    fc = {"oil": orch._place_curve(self_curve=_mk_curve("oil"),
                                   start_date="2026-01-01", strategy="pdp"),
          "gas": orch._place_curve(self_curve=_mk_curve("gas"),
                                   start_date="2026-01-01", strategy="pdp")}
    forecasts = {"42-111-11111": fc}
    classifications = {"42-111-11111": "history"}
    statuses = {"42-111-11111": "PRODUCING"}
    econ_overrides = {
        "interest_type": "minerals", "interest": {"decimal": 0.0625},
        "price_deck": {"type": "flat", "oil_usd_bbl": 72.0, "gas_usd_mmbtu": 3.1},
    }
    econ = orch._economics_from_forecasts(
        forecasts=forecasts, classifications=classifications,
        statuses=statuses, econ_overrides=econ_overrides)
    wells = {"statuses": statuses, "classifications": classifications}
    return econ, wells


def test_reconcile_total_pv_matches_engine():
    econ, wells = _engine_minerals_run()
    got = ex.reconcile_total_pv(econ, wells)
    want = econ["npv_at_centers"]["total"]
    assert abs(got - want) < 1.0, f"got {got}, want {want}"


def test_reconcile_total_pv_wi_with_costs():
    fc = {"oil": orch._place_curve(self_curve=_mk_curve("oil"),
                                   start_date="2026-01-01", strategy="pdp"),
          "gas": orch._place_curve(self_curve=_mk_curve("gas"),
                                   start_date="2026-01-01", strategy="pdp")}
    forecasts = {"42-222-22222": fc}
    classifications = {"42-222-22222": "history"}
    statuses = {"42-222-22222": "PRODUCING"}
    econ_overrides = {
        "interest_type": "wi", "interest": {"wi_pct": 0.75, "nri_pct": 0.5625},
        "price_deck": {"type": "flat", "oil_usd_bbl": 70.0, "gas_usd_mmbtu": 3.0},
        "opex_per_bbl_usd": 3.0, "opex_per_well_per_month_usd": 2000.0,
    }
    econ = orch._economics_from_forecasts(
        forecasts=forecasts, classifications=classifications,
        statuses=statuses, econ_overrides=econ_overrides)
    wells = {"statuses": statuses, "classifications": classifications}
    # Lock the override→cost_inputs plumbing: the costs must reach the oracle.
    a = ex.read_assumptions(econ)
    assert a["opex_bbl"] == pytest.approx(3.0)
    assert a["opex_well_month"] == pytest.approx(2000.0)
    assert abs(ex.reconcile_total_pv(econ, wells)
               - econ["npv_at_centers"]["total"]) < 1.0


def test_reconcile_total_pv_wi_capex():
    # A single no_history / PERMITTED (PUD) well books a drilling AFE at its
    # online month, exercising the oracle's capex = wi*capex_well*new_wells term.
    fc = {"oil": orch._place_curve(self_curve=_mk_curve("oil"),
                                   start_date="2027-01-01", strategy="pure_analog"),
          "gas": orch._place_curve(self_curve=_mk_curve("gas"),
                                   start_date="2027-01-01", strategy="pure_analog")}
    api = "42-333-33333"
    forecasts = {api: fc}
    classifications = {api: "no_history"}
    statuses = {api: "PERMITTED"}
    econ_overrides = {
        "interest_type": "wi", "interest": {"wi_pct": 0.75, "nri_pct": 0.5625},
        "price_deck": {"type": "flat", "oil_usd_bbl": 70.0, "gas_usd_mmbtu": 3.0},
        "capex_per_well_usd": 8_000_000.0,
    }
    econ = orch._economics_from_forecasts(
        forecasts=forecasts, classifications=classifications,
        statuses=statuses, econ_overrides=econ_overrides)
    wells = {"statuses": statuses, "classifications": classifications}
    # The AFE must actually book: at least one PUD-bucket month drills a well.
    drivers = ex.build_status_drivers(econ, wells)
    assert any(n > 0 for n in drivers["PUD"]["new_wells"])
    assert abs(ex.reconcile_total_pv(econ, wells)
               - econ["npv_at_centers"]["total"]) < 1.0


def _econ_two_wells():
    # 4-month horizon, two wells. Well A produces from month 0 (PDP),
    # well B comes online month 2 with a capex AFE (PUD/no_history).
    return {
        "horizon_months": 4,
        "schedule": {
            "origin": "2026-01-01",
            "by_well": {
                "A": {"oil_bbl": [100, 90, 80, 70], "gas_mcf": [10, 9, 8, 7],
                      "capex": [0, 0, 0, 0], "opex": [0, 0, 0, 0],
                      "net_oil": [0]*4, "net_gas": [0]*4, "gross_rev": [0]*4,
                      "net_rev": [0]*4, "sev_tax": [0]*4, "gpt": [0]*4,
                      "net_cashflow": [0]*4},
                "B": {"oil_bbl": [0, 0, 200, 180], "gas_mcf": [0, 0, 20, 18],
                      "capex": [0, 0, 5_000_000, 0], "opex": [0, 0, 0, 0],
                      "net_oil": [0]*4, "net_gas": [0]*4, "gross_rev": [0]*4,
                      "net_rev": [0]*4, "sev_tax": [0]*4, "gpt": [0]*4,
                      "net_cashflow": [0]*4},
            },
        },
    }


def _wells_two():
    return {"statuses": {"A": "PRODUCING", "B": "PERMITTED"},
            "classifications": {"A": "history", "B": "no_history"}}


def test_build_status_drivers_buckets_and_counts():
    drivers = ex.build_status_drivers(_econ_two_wells(), _wells_two())
    assert list(drivers.keys()) == ["PDP", "PUD"]    # no DUC; STATUS_ORDER preserved
    assert drivers["PDP"]["gross_oil"] == [100, 90, 80, 70]
    assert drivers["PDP"]["gross_gas"] == [10, 9, 8, 7]
    assert drivers["PDP"]["active_count"] == [1, 1, 1, 1]
    assert drivers["PDP"]["new_wells"] == [0, 0, 0, 0]
    assert drivers["PUD"]["gross_oil"] == [0, 0, 200, 180]
    assert drivers["PUD"]["gross_gas"] == [0, 0, 20, 18]
    assert drivers["PUD"]["active_count"] == [0, 0, 1, 1]
    assert drivers["PUD"]["new_wells"] == [0, 0, 1, 0]


def test_build_status_drivers_never_online_well():
    # A well that never produces (all-zero oil AND gas) and never drills:
    # _online_offset returns len(oil), so active_count stays 0 every month.
    econ = {
        "horizon_months": 4,
        "schedule": {
            "origin": "2026-01-01",
            "by_well": {
                "Z": {"oil_bbl": [0, 0, 0, 0], "gas_mcf": [0, 0, 0, 0],
                      "capex": [0, 0, 0, 0]},
            },
        },
    }
    wells = {"statuses": {"Z": "PRODUCING"}, "classifications": {"Z": "history"}}
    drivers = ex.build_status_drivers(econ, wells)
    assert drivers["PDP"]["active_count"] == [0, 0, 0, 0]
    assert drivers["PDP"]["new_wells"] == [0, 0, 0, 0]


def test_build_status_drivers_requires_by_well():
    econ = {"horizon_months": 4, "schedule": {"origin": "2026-01-01",
            "by_well_omitted": "too many wells"}}
    with pytest.raises(ex.ExportError):
        ex.build_status_drivers(econ, _wells_two())


# ── Task 4: workbook assembly ────────────────────────────────────────────────


def test_build_workbook_structure_and_formulas():
    from server.valuation import export_xlsx as ex
    econ, wells = _engine_minerals_run()
    facts = {"deal_type": "Minerals", "interest": "6.25% RI",
             "operator": "Test Op", "area": "Reeves County"}
    data = ex.build_workbook_bytes("run-abcdef12", econ, wells, facts)
    wb = load_workbook(io.BytesIO(data))                # formulas, not values
    # Two sheets only: Summary first, then the single aggregate Cashflow sheet.
    assert wb.sheetnames == ["Summary", "Cashflow"]
    assert "Assumptions" not in wb.sheetnames
    names = set(wb.defined_names)
    assert {"oil_diff", "gas_diff", "tax_pct"} <= names
    cf = wb["Cashflow"]
    f = ex._FIRST_DATA_ROW
    # No discounting anywhere in the workbook.
    flat = [c.value for row in cf.iter_rows() for c in row if isinstance(c.value, str)]
    assert not any("NPV(" in v for v in flat)
    # NetCF formula on the first data row mirrors the algebra; prices are inline.
    assert cf[f"N{f}"].value == f"=I{f}-J{f}-K{f}-L{f}-M{f}"
    assert cf[f"H{f}"].value == f"=B{f}*(F{f}-oil_diff)+C{f}*(G{f}-gas_diff)"
    assert cf[f"O{f}"].value == f"=N{f}"                       # cumulative seed
    assert cf[f"O{f + 1}"].value == f"=O{f}+N{f + 1}"          # running total
    # Summary's total references the Cashflow column, no PV line.
    summ_vals = [c.value for row in wb["Summary"].iter_rows()
                 for c in row if isinstance(c.value, str)]
    assert any("Total Net Cashflow" in v for v in summ_vals)
    assert not any("PV" == v or "Discount rate" in v for v in summ_vals)


def test_export_filename_slug():
    from server.valuation import export_xlsx as ex
    fn = ex.export_filename({"area": "Reeves County"}, "abcdef1234")
    assert fn.endswith(".xlsx") and "abcdef12" in fn


def test_build_workbook_wi_formulas():
    # The structure test only exercises the minerals branch; this one locks the
    # WI formula strings (nri/tax/gpt/opex/capex) + the WI-only named ranges.
    fc = {"oil": orch._place_curve(self_curve=_mk_curve("oil"),
                                   start_date="2026-01-01", strategy="pdp"),
          "gas": orch._place_curve(self_curve=_mk_curve("gas"),
                                   start_date="2026-01-01", strategy="pdp")}
    forecasts = {"42-222-22222": fc}
    classifications = {"42-222-22222": "history"}
    statuses = {"42-222-22222": "PRODUCING"}
    econ_overrides = {
        "interest_type": "wi", "interest": {"wi_pct": 0.75, "nri_pct": 0.5625},
        "price_deck": {"type": "flat", "oil_usd_bbl": 70.0, "gas_usd_mmbtu": 3.0},
        "opex_per_bbl_usd": 3.0, "opex_per_well_per_month_usd": 2000.0,
    }
    econ = orch._economics_from_forecasts(
        forecasts=forecasts, classifications=classifications,
        statuses=statuses, econ_overrides=econ_overrides)
    wells = {"statuses": statuses, "classifications": classifications}
    facts = {"deal_type": "Working Interest", "interest": "75% WI / 56.25% NRI",
             "operator": "Test Op", "area": "Reeves County"}
    data = ex.build_workbook_bytes("run-abcdef12", econ, wells, facts)
    wb = load_workbook(io.BytesIO(data))
    cf = wb["Cashflow"]
    f = ex._FIRST_DATA_ROW
    assert cf[f"I{f}"].value == f"=nri_pct*H{f}"
    assert cf[f"J{f}"].value == f"=tax_pct*wi_pct*H{f}"
    assert cf[f"K{f}"].value == f"=gpt_pct*wi_pct*H{f}"
    assert cf[f"L{f}"].value == f"=wi_pct*(opex_well*D{f}+opex_bbl*B{f})"
    assert cf[f"M{f}"].value == f"=wi_pct*capex_well*E{f}"
    names = set(wb.defined_names)
    assert {"wi_pct", "nri_pct"} <= names
    assert "dec_int" not in names


def test_aggregate_drivers_sums_across_statuses():
    # Two-well deal spanning PDP + PUD collapses to one aggregate volume column.
    econ, wells = _econ_two_wells(), _wells_two()
    drivers = ex.build_status_drivers(econ, wells)
    agg = ex._aggregate_drivers(drivers, econ["horizon_months"])
    # month 2: PDP well A (80) + PUD well B (200) = 280 gross oil; both active.
    assert agg["gross_oil"] == [100, 90, 280, 250]
    assert agg["active_count"] == [1, 1, 2, 2]
    assert agg["new_wells"] == [0, 0, 1, 0]


def test_workbook_table_rows_are_aggregated():
    # The single Cashflow sheet carries the summed volumes, not per-status splits.
    econ, wells = _econ_two_wells(), _wells_two()
    facts = {"deal_type": "Working Interest", "area": "Test"}
    wb = load_workbook(io.BytesIO(ex.build_workbook_bytes("run-x", econ, wells, facts)))
    assert wb.sheetnames == ["Summary", "Cashflow"]
    cf = wb["Cashflow"]
    f = ex._FIRST_DATA_ROW
    assert cf[f"B{f + 2}"].value == 280            # month-2 aggregate gross oil
    assert cf[f"D{f + 2}"].value == 2              # both wells active
