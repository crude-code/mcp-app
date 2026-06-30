import base64
import io
import json
from unittest.mock import patch

from openpyxl import load_workbook
import server.mcp_server as srv


def _fake_record():
    # Minimal economics + wells stages for a one-well minerals deal, 3 months.
    return {
        "run_id": "run-abcdef12",
        "user_id": 1,
        "economics": {
            "horizon_months": 3,
            "rate_centers": {"PDP": 0.10, "DUC": 0.20, "PUD": 0.25},
            "npv_at_centers": {"by_status": {"PDP": 0.0}, "total": 0.0},
            "interest": {"interest_type": "minerals", "decimal": 0.06},
            "inputs": {"oil_diff": 0.0, "gas_diff": 0.0, "tax_pct": 0.046, "gpt_pct": 0.0},
            "price_path": {"oil": [70, 70, 70], "gas": [3, 3, 3]},
            "cost_inputs": {"capex_per_well": 0.0, "opex_per_well_month": 0.0, "opex_per_bbl": 0.0},
            "schedule": {"origin": "2026-01-01", "by_well": {
                "A": {"oil_bbl": [100, 90, 80], "gas_mcf": [10, 9, 8], "capex": [0, 0, 0]},
            }},
        },
        "wells": {"well_meta": {"A": {"status": "PRODUCING", "operator": "Op",
                  "basin": "PB", "formation": "WC", "lateral_ft": 10000}},
                  "statuses": {"A": "PRODUCING"}, "classifications": {"A": "history"}},
    }


def test_export_valuation_xlsx_happy():
    rec = _fake_record()
    with patch.object(srv, "get_current_identity",
                      return_value={"user_slug": "u", "user_id": 1}), \
         patch.object(srv._valuation_store, "get", return_value=rec):
        out = json.loads(srv.export_valuation_xlsx("run-abcdef12"))
    assert "xlsx_base64" in out and out["filename"].endswith(".xlsx")
    wb = load_workbook(io.BytesIO(base64.b64decode(out["xlsx_base64"])))
    assert "Summary" in wb.sheetnames and "Cashflow" in wb.sheetnames


def test_export_valuation_xlsx_json_string_stages():
    # psycopg can return jsonb columns as raw JSON strings — exercise the
    # isinstance(..., str) parse guard in the tool.
    rec = _fake_record()
    rec["economics"] = json.dumps(rec["economics"])
    rec["wells"] = json.dumps(rec["wells"])
    with patch.object(srv, "get_current_identity",
                      return_value={"user_slug": "u", "user_id": 1}), \
         patch.object(srv._valuation_store, "get", return_value=rec):
        out = json.loads(srv.export_valuation_xlsx("run-abcdef12"))
    assert "xlsx_base64" in out and out["filename"].endswith(".xlsx")
    wb = load_workbook(io.BytesIO(base64.b64decode(out["xlsx_base64"])))
    assert "Summary" in wb.sheetnames


def test_export_valuation_xlsx_unknown_run():
    with patch.object(srv, "get_current_identity",
                      return_value={"user_slug": "u", "user_id": 1}), \
         patch.object(srv._valuation_store, "get", return_value=None):
        out = json.loads(srv.export_valuation_xlsx("nope"))
    assert "error" in out
